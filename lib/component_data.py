"""
Component library data structures and database
Now loads data from CSV files in the assets folder
"""

import os
import pandas as pd
from dataclasses import dataclass
from typing import List

@dataclass
class MOSFET:
    """MOSFET component specification with extended heuristics parameters"""
    name: str
    manufacturer: str
    vds: float  # Drain-Source Voltage (V)
    id: float  # Continuous Drain Current (A)
    rdson: float  # On-Resistance (mΩ)
    qg: float  # Total Gate Charge (nC), use 0 for N/A
    package: str
    typical_use: str = ""  # e.g., "12−24V", can be empty
    efficiency_range: str = ""  # e.g., "96–98%", can be empty
    high_side_ok: bool = True  # Can be used as high-side switch
    low_side_ok: bool = True  # Can be used as low-side switch
    voltage_domain: str = ""  # e.g., "12−24V Robotics"
    # NEW: Extended heuristics parameters (from MOSFET Design heuristics)
    vgs_max: float = 20.0  # Max Gate-Source Voltage (V) - 20V for Si, 18V for SiC
    qgd: float = 0.0  # Gate-Drain Charge (nC) for dv/dt immunity calculation
    qgs: float = 0.0  # Gate-Source Charge (nC) for dv/dt immunity calculation
    package_inductance: float = 0.0  # Package inductance (nH) for dv/dt susceptibility
    junction_temp_max: float = 150.0  # Max junction temperature (°C)
    rdson_at_125c: float = 0.0  # RDS(on) at 125°C for temperature derating (mΩ)
    mosfet_type: str = "Si"  # Si or SiC - affects VDS derating factors

@dataclass
class Capacitor:
    """Output Capacitor component specification"""
    part_number: str
    manufacturer: str
    capacitance: float  # µF
    voltage: float  # V
    type: str
    esr: str  # mΩ
    primary_use: str
    temp_range: str

@dataclass
class InputCapacitor:
    """Input Capacitor component specification with ripple current handling"""
    part_number: str
    manufacturer: str
    category: str  # MLCC, Polymer, Electrolytic, Film
    dielectric: str  # X7R, X5R, Conductive Polymer, etc.
    capacitance: float  # µF
    voltage: float  # V
    esr: float  # mΩ
    esl: float  # nH (Equivalent Series Inductance)
    ripple_rating: float  # A (if available)
    lifetime: float  # hours (if available)
    package: str
    cost: float  # USD (if available)
    availability: str
    notes: str

@dataclass
class Inductor:
    """Inductor component specification"""
    part_number: str
    manufacturer: str
    inductance: float  # µH
    current: float  # A
    dcr: float  # mΩ (DC Resistance)
    sat_current: float  # A
    package: str
    shielded: bool = False  # New field
    core_material: str = ""  # New field
    temp_range: str = ""  # New field


def load_mosfets_from_excel() -> List[MOSFET]:
    """Load MOSFETs from PowerCrux Excel file (PRIMARY SOURCE)"""
    try:
        import openpyxl
        
        current_dir = os.path.dirname(os.path.abspath(__file__))
        
        possible_paths = [
            os.path.join(current_dir, '..', 'assets', 'component_data', 'PowerCrux_SAFE_LINKS_EXACT_50_Si_MOSFETs_SyncBuck.xlsx'),
            os.path.join(os.getcwd(), 'assets', 'component_data', 'PowerCrux_SAFE_LINKS_EXACT_50_Si_MOSFETs_SyncBuck.xlsx'),
            'assets/component_data/PowerCrux_SAFE_LINKS_EXACT_50_Si_MOSFETs_SyncBuck.xlsx'
        ]
        
        excel_path = None
        for path in possible_paths:
            if os.path.exists(path):
                excel_path = path
                break
        
        if not excel_path:
            print(f"Info: PowerCrux MOSFET Excel file not found, falling back to CSV")
            return load_mosfets_from_csv()
        
        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active
        mosfets = []
        
        # Skip header row
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 1):
            if not row[0]:  # Skip empty rows
                continue
            try:
                # Map Excel columns: Manufacturer, Part Number, Vds, Id, Rds_on_mOhm, Qg_nC, Package, 
                #                    High_Side_OK, Low_Side_OK, Robotics_Voltage_Domain, Datasheet_Search_Link
                mosfet = MOSFET(
                    manufacturer=str(row[0]) if row[0] else "Unknown",
                    name=str(row[1]) if row[1] else "Unknown",
                    vds=float(row[2]) if row[2] else 0.0,
                    id=float(row[3]) if row[3] else 0.0,
                    rdson=float(row[4]) if row[4] else 0.0,
                    qg=float(row[5]) if row[5] else 0.0,
                    package=str(row[6]) if row[6] else "Unknown",
                    high_side_ok=str(row[7]).lower() == 'yes' if row[7] else True,
                    low_side_ok=str(row[8]).lower() == 'yes' if row[8] else True,
                    voltage_domain=str(row[9]) if row[9] else "",
                    typical_use="PowerCrux optimized",
                    efficiency_range="High efficiency"
                )
                mosfets.append(mosfet)
            except Exception as e:
                print(f"Warning: Failed to parse MOSFET row {row_idx}: {e}")
                continue
        
        print(f"Loaded {len(mosfets)} MOSFETs from PowerCrux Excel file")
        return mosfets if mosfets else load_mosfets_from_csv()
    
    except ImportError:
        print("Warning: openpyxl not installed, falling back to CSV loader")
        return load_mosfets_from_csv()
    except Exception as e:
        print(f"Error loading MOSFET data from Excel: {e}")
        return load_mosfets_from_csv()


def load_inductors_from_excel() -> List[Inductor]:
    """Load Inductors from PowerCrux Excel file (PRIMARY SOURCE)"""
    try:
        import openpyxl
        
        current_dir = os.path.dirname(os.path.abspath(__file__))
        
        possible_paths = [
            os.path.join(current_dir, '..', 'assets', 'component_data', 'powercrux_inductors_20parts (2).xlsx'),
            os.path.join(os.getcwd(), 'assets', 'component_data', 'powercrux_inductors_20parts (2).xlsx'),
            'assets/component_data/powercrux_inductors_20parts (2).xlsx'
        ]
        
        excel_path = None
        for path in possible_paths:
            if os.path.exists(path):
                excel_path = path
                break
        
        if not excel_path:
            print(f"Info: PowerCrux Inductor Excel file not found, falling back to CSV")
            return load_inductors_from_csv()
        
        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active
        inductors = []
        
        # Skip header row
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 1):
            if not row[0]:  # Skip empty rows
                continue
            try:
                # Columns: manufacturer, part_number, L_nom_uH, tolerance, DCR_mOhm_max, I_rated_A, 
                #          I_sat_A, package, shielded, core_material, operating_temp_C, notes
                inductor = Inductor(
                    manufacturer=str(row[0]) if row[0] else "Unknown",
                    part_number=str(row[1]) if row[1] else "Unknown",
                    inductance=float(row[2]) if row[2] else 0.0,
                    dcr=float(row[4]) if row[4] else 0.0,
                    current=float(row[5]) if row[5] else 0.0,
                    sat_current=float(row[6]) if row[6] else 0.0,
                    package=str(row[7]) if row[7] else "Unknown",
                    shielded=str(row[8]).lower() == 'true' if row[8] else False,
                    core_material=str(row[9]) if row[9] else "",
                    temp_range=str(row[10]) if row[10] else ""
                )
                inductors.append(inductor)
            except Exception as e:
                print(f"Warning: Failed to parse Inductor row {row_idx}: {e}")
                continue
        
        print(f"Loaded {len(inductors)} Inductors from PowerCrux Excel file")
        return inductors if inductors else load_inductors_from_csv()
    
    except ImportError:
        print("Warning: openpyxl not installed, falling back to CSV loader")
        return load_inductors_from_csv()
    except Exception as e:
        print(f"Error loading Inductor data from Excel: {e}")
        return load_inductors_from_csv()


def load_input_capacitors_from_excel() -> List[InputCapacitor]:
    """Load Input Capacitors from PowerCrux Excel file (PRIMARY SOURCE)"""
    try:
        import openpyxl
        
        current_dir = os.path.dirname(os.path.abspath(__file__))
        
        possible_paths = [
            os.path.join(current_dir, '..', 'assets', 'component_data', 'input_capacitor_db_powercrux.xlsx'),
            os.path.join(os.getcwd(), 'assets', 'component_data', 'input_capacitor_db_powercrux.xlsx'),
            'assets/component_data/input_capacitor_db_powercrux.xlsx'
        ]
        
        excel_path = None
        for path in possible_paths:
            if os.path.exists(path):
                excel_path = path
                break
        
        if not excel_path:
            print(f"Info: PowerCrux Input Capacitor Excel file not found, falling back to CSV")
            return load_input_capacitors_from_csv()
        
        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active
        capacitors = []
        
        # Skip header row
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 1):
            if not row[0]:  # Skip empty rows
                continue
            try:
                # Columns: Part Number, Manufacturer, Category, Dielectric, Rated Capacitance (uF),
                #          Rated Voltage (V), ESR (mOhm), ESL (nH), Ripple Rating (A), 
                #          Lifetime (h), Package, Cost (USD), Availability, Notes
                capacitor = InputCapacitor(
                    part_number=str(row[0]) if row[0] else "Unknown",
                    manufacturer=str(row[1]) if row[1] else "Unknown",
                    category=str(row[2]) if row[2] else "Unknown",
                    dielectric=str(row[3]) if row[3] else "Unknown",
                    capacitance=float(row[4]) if row[4] else 0.0,
                    voltage=float(row[5]) if row[5] else 0.0,
                    esr=float(row[6]) if row[6] else 0.0,
                    esl=float(row[7]) if row[7] else 0.0,
                    ripple_rating=float(row[8]) if row[8] else 0.0,
                    lifetime=float(row[9]) if row[9] else 0.0,
                    package=str(row[10]) if row[10] else "Unknown",
                    cost=float(row[11]) if row[11] else 0.0,
                    availability=str(row[12]) if row[12] else "Unknown",
                    notes=str(row[13]) if row[13] else ""
                )
                capacitors.append(capacitor)
            except Exception as e:
                print(f"Warning: Failed to parse Input Capacitor row {row_idx}: {e}")
                continue
        
        print(f"Loaded {len(capacitors)} Input Capacitors from PowerCrux Excel file")
        return capacitors if capacitors else load_input_capacitors_from_csv()
    
    except ImportError:
        print("Warning: openpyxl not installed, falling back to CSV loader")
        return load_input_capacitors_from_csv()
    except Exception as e:
        print(f"Error loading Input Capacitor data from Excel: {e}")
        return load_input_capacitors_from_csv()


def load_mosfets_from_csv() -> List[MOSFET]:
    """Load MOSFETs from CSV file with robust path handling (FALLBACK)"""
    try:
        # Multiple path resolution strategies for different environments
        current_dir = os.path.dirname(os.path.abspath(__file__))
        
        # Try different path combinations
        possible_paths = [
            os.path.join(current_dir, '..', 'assets', 'component_data', 'mosfets.csv'),
            os.path.join(os.getcwd(), 'assets', 'component_data', 'mosfets.csv'),
            os.path.join(current_dir, '..', '..', 'assets', 'component_data', 'mosfets.csv'),
            'assets/component_data/mosfets.csv'  # Relative path for cloud deployment
        ]
        
        csv_path = None
        for path in possible_paths:
            if os.path.exists(path):
                csv_path = path
                break
        
        if not csv_path:
            print(f"Warning: MOSFET CSV file not found in any of these locations: {possible_paths}")
            return get_fallback_mosfets()
        
        df = pd.read_csv(csv_path)
        mosfets = []
        
        for _, row in df.iterrows():
            mosfet = MOSFET(
                name=str(row['name']),
                manufacturer=str(row['manufacturer']),
                vds=float(row['vds']),
                id=float(row['id']),
                rdson=float(row['rdson']),
                qg=float(row['qg']),
                package=str(row['package']),
                typical_use=str(row['typical_use']),
                efficiency_range=str(row['efficiency_range'])
            )
            mosfets.append(mosfet)
        
        return mosfets
    except Exception as e:
        print(f"Error loading MOSFET data from CSV: {e}")
        return get_fallback_mosfets()


def load_capacitors_from_csv() -> List[Capacitor]:
    """Load Capacitor data from CSV file with robust path handling (FALLBACK)"""
    try:
        # Multiple path resolution strategies for different environments
        current_dir = os.path.dirname(os.path.abspath(__file__))
        
        # Try different path combinations for output capacitors
        possible_paths = [
            os.path.join(current_dir, '..', 'assets', 'component_data', 'output_capacitors.csv'),
            os.path.join(os.getcwd(), 'assets', 'component_data', 'output_capacitors.csv'),
            os.path.join(current_dir, '..', '..', 'assets', 'component_data', 'output_capacitors.csv'),
            'assets/component_data/output_capacitors.csv'  # Relative path for cloud deployment
        ]
        
        csv_path = None
        for path in possible_paths:
            if os.path.exists(path):
                csv_path = path
                break
        
        if not csv_path:
            print(f"Warning: Capacitor CSV file not found in any of these locations: {possible_paths}")
            return get_fallback_capacitors()
        
        df = pd.read_csv(csv_path)
        capacitors = []
        
        for _, row in df.iterrows():
            capacitor = Capacitor(
                part_number=str(row['part_number']),
                manufacturer=str(row['manufacturer']),
                capacitance=float(row['capacitance']),
                voltage=float(row['voltage']),
                type=str(row['type']),
                esr=str(row['esr']),
                primary_use=str(row['primary_use']),
                temp_range=str(row['temp_range'])
            )
            capacitors.append(capacitor)
        
        return capacitors
    except Exception as e:
        print(f"Error loading Capacitor data from CSV: {e}")
        return get_fallback_capacitors()


def load_input_capacitors_from_csv() -> List[InputCapacitor]:
    """Load Input Capacitor data from CSV file with robust path handling (FALLBACK)"""
    try:
        # Multiple path resolution strategies for different environments
        current_dir = os.path.dirname(os.path.abspath(__file__))
        
        # Try different path combinations for input capacitors
        possible_paths = [
            os.path.join(current_dir, '..', 'assets', 'component_data', 'input_capacitors.csv'),
            os.path.join(os.getcwd(), 'assets', 'component_data', 'input_capacitors.csv'),
            os.path.join(current_dir, '..', '..', 'assets', 'component_data', 'input_capacitors.csv'),
            'assets/component_data/input_capacitors.csv'  # Relative path for cloud deployment
        ]
        
        csv_path = None
        for path in possible_paths:
            if os.path.exists(path):
                csv_path = path
                break
        
        if not csv_path:
            print(f"Warning: Input Capacitor CSV file not found in any of these locations: {possible_paths}")
            return []  # Return empty list if file not found
        
        df = pd.read_csv(csv_path)
        input_capacitors = []
        
        for _, row in df.iterrows():
            input_capacitor = InputCapacitor(
                part_number=str(row['part_number']),
                manufacturer=str(row['manufacturer']),
                category=str(row['category']),
                dielectric=str(row['dielectric']),
                capacitance=float(row['rated_capacitance_uf']),
                voltage=float(row['rated_voltage_v']),
                esr=float(row['esr_mohm']) if pd.notna(row['esr_mohm']) else 0.0,
                esl=float(row['esl_nh']) if pd.notna(row['esl_nh']) else 0.0,
                ripple_rating=float(row['ripple_rating_a']) if pd.notna(row['ripple_rating_a']) else 0.0,
                lifetime=float(row['lifetime_h']) if pd.notna(row['lifetime_h']) else 0.0,
                package=str(row['package']),
                cost=float(row['cost_usd']) if pd.notna(row['cost_usd']) else 0.0,
                availability=str(row['availability']),
                notes=str(row['notes'])
            )
            input_capacitors.append(input_capacitor)
        
        return input_capacitors
    except Exception as e:
        print(f"Error loading Input Capacitor data from CSV: {e}")
        return []


def load_inductors_from_csv() -> List[Inductor]:
    """Load Inductor data from CSV file with robust path handling (FALLBACK)"""
    try:
        # Multiple path resolution strategies for different environments
        current_dir = os.path.dirname(os.path.abspath(__file__))
        
        # Try different path combinations
        possible_paths = [
            os.path.join(current_dir, '..', 'assets', 'component_data', 'inductors.csv'),
            os.path.join(os.getcwd(), 'assets', 'component_data', 'inductors.csv'),
            os.path.join(current_dir, '..', '..', 'assets', 'component_data', 'inductors.csv'),
            'assets/component_data/inductors.csv'  # Relative path for cloud deployment
        ]
        
        csv_path = None
        for path in possible_paths:
            if os.path.exists(path):
                csv_path = path
                break
        
        if not csv_path:
            print(f"Warning: Inductor CSV file not found in any of these locations: {possible_paths}")
            return get_fallback_inductors()
        
        df = pd.read_csv(csv_path)
        inductors = []
        
        for _, row in df.iterrows():
            # Handle both old and new CSV format - check for new format columns first
            if 'l_nom_uh' in df.columns:
                # New PowerCrux format
                inductor = Inductor(
                    part_number=str(row['part_number']),
                    manufacturer=str(row['manufacturer']),
                    inductance=float(row['l_nom_uh']),
                    current=float(row['i_rated_a']),
                    dcr=float(row['dcr_mohm_max']),
                    sat_current=float(row['i_sat_a']),
                    package=str(row['package']),
                    shielded=str(row['shielded']).lower() == 'true' if 'shielded' in df.columns else False,
                    core_material=str(row['core_material']) if 'core_material' in df.columns else "",
                    temp_range=str(row['operating_temp_c']) if 'operating_temp_c' in df.columns else ""
                )
            else:
                # Old format
                inductor = Inductor(
                    part_number=str(row['part_number']),
                    manufacturer=str(row['manufacturer']),
                    inductance=float(row['inductance']),
                    current=float(row['current']),
                    dcr=float(row['dcr']),
                    sat_current=float(row['sat_current']),
                    package=str(row['package'])
                )
            inductors.append(inductor)
        
        return inductors
    except Exception as e:
        print(f"Error loading Inductor data from CSV: {e}")
        return get_fallback_inductors()


def get_fallback_mosfets() -> List[MOSFET]:
    """Fallback MOSFET data if Excel/CSV loading fails"""
    return [
        MOSFET("BSC016N06NS", "Infineon", 60, 150, 1.6, 44, "SuperSO8", 
               typical_use="Mid-power buck, low loss", efficiency_range="96–98%", 
               high_side_ok=True, low_side_ok=True, voltage_domain="12−24V"),
        MOSFET("CSD19505KCS", "Texas Instruments", 60, 80, 2.5, 37, "TO-220",
               typical_use="Classic synchronous buck", efficiency_range="95–97%",
               high_side_ok=True, low_side_ok=True, voltage_domain="12−24V"),
        MOSFET("IPB017N10N5", "Infineon", 100, 120, 1.7, 70, "D²PAK",
               typical_use="48V bus robotics converter", efficiency_range="96–97%",
               high_side_ok=True, low_side_ok=True, voltage_domain="24−48V"),
    ]


def get_fallback_capacitors() -> List[Capacitor]:
    """Fallback Capacitor data if CSV loading fails"""
    return [
        Capacitor("C3216X7R1H106K160AC", "TDK", 10, 50, "MLCC X7R", "~2-5", "HF ripple + local decoupling", "-55..125"),
        Capacitor("C3225X7R1E226M250AB", "TDK", 22, 25, "MLCC X7R", "~2-5", "HF ripple + bulk on 12V", "-55..125"),
    ]


def get_fallback_inductors() -> List[Inductor]:
    """Fallback Inductor data if CSV loading fails"""
    return [
        Inductor("SER2915H-472KL", "Coilcraft", 4700, 0.73, 1850, 0.95, "SER2915"),
        Inductor("SER2915H-103KL", "Coilcraft", 10000, 0.48, 4200, 0.62, "SER2915"),
    ]


# Load data from Excel files (PRIMARY) with CSV fallback
MOSFET_LIBRARY: List[MOSFET] = load_mosfets_from_excel()
INDUCTOR_LIBRARY: List[Inductor] = load_inductors_from_excel()
INPUT_CAPACITOR_LIBRARY: List[InputCapacitor] = load_input_capacitors_from_excel()
CAPACITOR_LIBRARY: List[Capacitor] = load_capacitors_from_csv()  # Output capacitors (CSV only)


def reload_component_data():
    """Reload all component data from Excel files (with CSV fallback)"""
    global MOSFET_LIBRARY, CAPACITOR_LIBRARY, INDUCTOR_LIBRARY, INPUT_CAPACITOR_LIBRARY
    MOSFET_LIBRARY = load_mosfets_from_excel()
    INDUCTOR_LIBRARY = load_inductors_from_excel()
    INPUT_CAPACITOR_LIBRARY = load_input_capacitors_from_excel()
    CAPACITOR_LIBRARY = load_capacitors_from_csv()
    print("Component data reloaded from Excel files (with CSV fallback)")


def get_design_heuristics_path() -> str:
    """Get the path to the design heuristics folder"""
    current_dir = os.path.dirname(os.path.abspath(__file__))
    return os.path.join(current_dir, '..', 'assets', 'design_heuristics')
